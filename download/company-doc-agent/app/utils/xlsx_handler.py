"""
XLSX 表格处理工具模块
负责：读取 XLSX 内容为文本 / 将文本/结构化数据写回 XLSX

功能：
1. read_xlsx_to_text() — 读取 XLSX 文件内容为结构化文本（供 RAG 索引和 LLM 理解）
2. write_xlsx_from_text() — 将 LLM 修改后的文本写回 XLSX 文件
3. parse_table_text() — 解析 LLM 输出的表格文本为行列数据

依赖：openpyxl
"""
import os
import re
import logging
from typing import Optional

logger = logging.getLogger(__name__)


def read_xlsx_to_text(file_path: str, max_rows: int = 500, max_cols: int = 50) -> str:
    """
    读取 XLSX 文件内容，转换为结构化文本

    每个 Sheet 输出格式：
    === Sheet: 工作表名 ===
    | 列1 | 列2 | 列3 |
    |------|------|------|
    | 值1  | 值2  | 值3  |

    Args:
        file_path: XLSX 文件路径
        max_rows: 每个 Sheet 最大读取行数
        max_cols: 每个 Sheet 最大读取列数

    Returns:
        str: 结构化文本内容
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    output_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output_parts.append(f"=== Sheet: {sheet_name} ===")

        rows_data = []
        row_count = 0

        for row in ws.iter_rows(max_col=max_cols, max_row=max_rows, values_only=False):
            if row_count >= max_rows:
                output_parts.append(f"... (超过 {max_rows} 行，已截断)")
                break

            row_values = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_values.append("")
                else:
                    row_values.append(str(val))
            rows_data.append(row_values)
            row_count += 1

        if not rows_data:
            output_parts.append("（空工作表）\n")
            continue

        # 格式化为 Markdown 表格
        # 确定列数（取最大列数）
        max_col_count = max(len(r) for r in rows_data) if rows_data else 0

        # 补齐每行的列数
        for row in rows_data:
            while len(row) < max_col_count:
                row.append("")

        # 输出表格
        for i, row in enumerate(rows_data):
            line = "| " + " | ".join(str(v).replace("|", "｜") for v in row) + " |"
            output_parts.append(line)
            # 第一行后加分隔线（表头分隔）
            if i == 0:
                separator = "|" + "|".join(["------" for _ in row]) + "|"
                output_parts.append(separator)

        output_parts.append("")  # Sheet 间空行

    wb.close()

    return "\n".join(output_parts)


def parse_table_text(text: str) -> list[list[list[str]]]:
    """
    解析 LLM 输出的表格文本为结构化数据（支持多个 Sheet）

    支持的格式：
    1. Markdown 表格格式：| col1 | col2 |
    2. 制表符分隔格式：col1\tcol2\tcol3
    3. Sheet 标记：=== Sheet: 名称 ===

    Args:
        text: LLM 输出的表格文本

    Returns:
        list[list[list[str]]]: [sheet1_rows, sheet2_rows, ...]
        每个 sheet 是 [[cell, cell, ...], [cell, cell, ...], ...]
    """
    # 按 Sheet 标记拆分
    sheet_pattern = re.compile(r'===\s*Sheet[:：]\s*(.+?)\s*===')
    lines = text.strip().split("\n")

    # 收集所有 sheet
    sheets = []  # [(sheet_name, [lines])]
    current_sheet_name = "Sheet1"
    current_lines = []

    for line in lines:
        match = sheet_pattern.search(line)
        if match:
            # 保存前一个 sheet
            if current_lines:
                sheets.append((current_sheet_name, current_lines))
            current_sheet_name = match.group(1).strip()
            current_lines = []
        else:
            current_lines.append(line)

    # 保存最后一个 sheet
    if current_lines:
        sheets.append((current_sheet_name, current_lines))

    # 如果没有 Sheet 标记，把全部内容当作一个 sheet
    if not sheets:
        sheets = [("Sheet1", lines)]

    # 解析每个 sheet 的行
    result = []
    for sheet_name, sheet_lines in sheets:
        rows = _parse_sheet_lines(sheet_lines)
        if rows:
            result.append(rows)

    return result


def _parse_sheet_lines(lines: list[str]) -> list[list[str]]:
    """解析单个 sheet 的文本行为表格行"""
    rows = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 跳过分隔行：|------|------|
        if re.match(r'^[\s|\-:]+$', line):
            continue

        # Markdown 表格格式：| val | val |
        if line.startswith("|") and line.endswith("|"):
            cells = [c.strip() for c in line[1:-1].split("|")]
            # 过滤全空行
            if any(c for c in cells):
                rows.append(cells)
            continue

        # 也处理 | val | val | 格式但末尾无 |
        if line.startswith("|"):
            cells = [c.strip() for c in line[1:].split("|")]
            if any(c for c in cells):
                rows.append(cells)
            continue

        # 制表符分隔格式
        if "\t" in line:
            cells = line.split("\t")
            if any(c.strip() for c in cells):
                rows.append([c.strip() for c in cells])
            continue

        # 逗号分隔（简单处理，不考虑引号内的逗号）
        if "," in line and not any(c.isalpha() and ord(c) > 127 for c in line.split(",")[0]):
            # 仅当第一列不含中文时才认为是 CSV 格式
            pass  # 不强制 CSV 解析，避免误判

        # 普通文本行 — 作为单单元格行
        rows.append([line])

    return rows


def write_xlsx_from_text(text: str, output_path: str, source_file: Optional[str] = None) -> str:
    """
    将 LLM 修改后的文本内容写入 XLSX 文件

    支持的输入格式：
    - Markdown 表格：| col1 | col2 |
    - 制表符分隔：col1\tcol2
    - 混合格式（带 Sheet 标记）

    Args:
        text: LLM 输出的表格文本
        output_path: 输出 XLSX 文件路径
        source_file: 原始 XLSX 文件路径（可选，用于复制样式）

    Returns:
        str: 实际输出文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    sheets_data = parse_table_text(text)

    if not sheets_data:
        # 如果解析失败，把整段文本写入一个单元格
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value=text)
        wb.save(output_path)
        return output_path

    wb = Workbook()
    # 删除默认创建的 Sheet（后面会重新添加）
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 解析 Sheet 名称
    sheet_pattern = re.compile(r'===\s*Sheet[:：]\s*(.+?)\s*===')
    lines = text.strip().split("\n")
    sheet_names = []
    for line in lines:
        match = sheet_pattern.search(line)
        if match:
            sheet_names.append(match.group(1).strip())

    for i, rows in enumerate(sheets_data):
        # 获取 Sheet 名称
        sheet_name = sheet_names[i] if i < len(sheet_names) else f"Sheet{i + 1}"
        # Sheet 名称最长 31 字符（Excel 限制）
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = cell_alignment
                cell.border = thin_border

                # 第一行作为表头加粗
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # 自动调整列宽（根据内容长度估算）
        for col_idx in range(1, max(len(r) for r in rows) + 1 if rows else 1):
            max_length = 0
            for row in rows:
                if col_idx <= len(row):
                    # 中文字符占 2 个宽度
                    val = str(row[col_idx - 1])
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_length = max(max_length, length)
            # 设置列宽（最小 8，最大 50）
            adjusted_width = min(max(max_length + 2, 8), 50)
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)
    logger.info(f"XLSX 文件生成成功: {output_path}")
    return output_path


def read_xlsx_content_simple(file_path: str) -> str:
    """
    简化版 XLSX 读取（用于 RAG 索引场景）
    将表格内容转为纯文本，每个单元格用空格分隔

    Args:
        file_path: XLSX 文件路径

    Returns:
        str: 纯文本内容
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    output_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output_parts.append(f"[工作表: {sheet_name}]")

        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                output_parts.append(row_text)

        output_parts.append("")

    wb.close()
    return "\n".join(output_parts)
